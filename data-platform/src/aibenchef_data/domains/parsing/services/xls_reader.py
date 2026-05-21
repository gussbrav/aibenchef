"""XlsReader — interfaz unificada para leer .xls SBS sin importar el formato real.

SBS publica bajo extension .xls al menos 5 formatos distintos:
- BIFF (xlrd)
- XLSX (openpyxl)
- XLSB (pyxlsb)
- HTML disfrazado (BeautifulSoup + lxml)
- SpreadsheetML XML 2003 (lxml)
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Protocol

from aibenchef_data.domains.shared import ValidationError

from ..value_objects.xls_format import XlsFormat, detect_xls_format

Cell = str | int | float | bool | None


@dataclass(frozen=True, slots=True)
class XlsSheet:
    name: str
    n_rows: int
    n_cols: int
    rows: list[list[Cell]]

    def row(self, idx: int) -> list[Cell]:
        return self.rows[idx]

    def cell(self, r: int, c: int) -> Cell:
        if r >= self.n_rows or c >= len(self.rows[r]):
            return None
        return self.rows[r][c]


class XlsReader(Protocol):
    def read(self, path: Path) -> list[XlsSheet]: ...


def read_xls(path: Path) -> list[XlsSheet]:
    """Lee cualquier .xls SBS detectando el formato y ruteando al lector correcto."""
    if not path.exists():
        raise ValidationError(f"Archivo no existe: {path}")

    fmt = detect_xls_format(path)

    if fmt == XlsFormat.BIFF:
        return _read_biff(path)
    if fmt == XlsFormat.XLSX:
        return _read_xlsx(path)
    if fmt == XlsFormat.XLSB:
        return _read_xlsb(path)
    if fmt == XlsFormat.HTML:
        return _read_html(path)
    if fmt == XlsFormat.XML2003:
        return _read_xml2003(path)

    raise ValidationError(f"Formato no soportado para {path.name}", context={"format": fmt.value})


# ============================================================================
# BIFF
# ============================================================================


def _read_biff(path: Path) -> list[XlsSheet]:
    import xlrd

    book = xlrd.open_workbook(str(path), formatting_info=False)
    sheets: list[XlsSheet] = []
    for sheet in book.sheets():
        rows: list[list[Cell]] = []
        for r in range(sheet.nrows):
            row_cells: list[Cell] = []
            for c in range(sheet.ncols):
                row_cells.append(_normalize_xlrd_cell(sheet, r, c))
            rows.append(row_cells)
        sheets.append(XlsSheet(name=sheet.name, n_rows=sheet.nrows, n_cols=sheet.ncols, rows=rows))
    return sheets


def _normalize_xlrd_cell(sheet: Any, r: int, c: int) -> Cell:
    import xlrd

    t = sheet.cell_type(r, c)
    v = sheet.cell_value(r, c)
    if t == xlrd.XL_CELL_EMPTY or t == xlrd.XL_CELL_BLANK:
        return None
    if t == xlrd.XL_CELL_TEXT:
        s = str(v).strip()
        return s or None
    if t == xlrd.XL_CELL_NUMBER:
        return v
    if t == xlrd.XL_CELL_BOOLEAN:
        return bool(v)
    if t == xlrd.XL_CELL_ERROR:
        return None
    return v


# ============================================================================
# XLSX (OOXML)
# ============================================================================


def _read_xlsx(path: Path) -> list[XlsSheet]:
    """Lee XLSX. Si el archivo tiene extension .xls (SBS publica algunos asi),
    openpyxl lo rechaza por el nombre aunque internamente sea zip-based.
    Workaround: pasarlo como BytesIO (sin nombre de archivo)."""
    from io import BytesIO

    from openpyxl import load_workbook

    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    else:
        # Cargar via BytesIO para que openpyxl no mire la extension
        buf = BytesIO(path.read_bytes())
        wb = load_workbook(filename=buf, read_only=True, data_only=True)

    try:
        sheets: list[XlsSheet] = []
        for ws in wb.worksheets:
            rows: list[list[Cell]] = []
            max_cols = 0
            for row in ws.iter_rows(values_only=True):
                row_cells: list[Cell] = [_normalize(v) for v in row]
                rows.append(row_cells)
                max_cols = max(max_cols, len(row_cells))
            sheets.append(XlsSheet(name=ws.title, n_rows=len(rows), n_cols=max_cols, rows=rows))
        return sheets
    finally:
        wb.close()


# ============================================================================
# XLSB (Excel Binary)
# ============================================================================


def _read_xlsb(path: Path) -> list[XlsSheet]:
    from pyxlsb import open_workbook

    sheets: list[XlsSheet] = []
    with open_workbook(str(path)) as wb:
        for sheet_name in wb.sheets:
            with wb.get_sheet(sheet_name) as ws:
                rows: list[list[Cell]] = []
                max_cols = 0
                for row in ws.rows():
                    cells: list[Cell] = [_normalize(cell.v) for cell in row]
                    rows.append(cells)
                    max_cols = max(max_cols, len(cells))
                sheets.append(
                    XlsSheet(
                        name=str(sheet_name),
                        n_rows=len(rows),
                        n_cols=max_cols,
                        rows=rows,
                    )
                )
    return sheets


# ============================================================================
# HTML disfrazado de .xls
# ============================================================================


def _read_html(path: Path) -> list[XlsSheet]:
    from bs4 import BeautifulSoup

    raw = path.read_bytes()
    text = _decode_html(raw)
    soup = BeautifulSoup(text, "lxml")

    sheets: list[XlsSheet] = []
    for idx, table in enumerate(soup.find_all("table")):
        rows: list[list[Cell]] = []
        max_cols = 0
        for tr in table.find_all("tr"):
            cells: list[Cell] = []
            for td in tr.find_all(["td", "th"]):
                txt = td.get_text(strip=True)
                cells.append(_parse_html_cell(txt))
            rows.append(cells)
            max_cols = max(max_cols, len(cells))
        sheets.append(XlsSheet(name=f"table_{idx}", n_rows=len(rows), n_cols=max_cols, rows=rows))
    return sheets


def _decode_html(raw: bytes) -> str:
    for enc in ("utf-8", "windows-1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace")


def _parse_html_cell(txt: str) -> Cell:
    if not txt:
        return None
    cleaned = txt.replace(",", "").replace("\xa0", "").strip()
    if not cleaned:
        return None
    try:
        if "." not in cleaned:
            return int(cleaned)
    except ValueError:
        pass
    try:
        return float(cleaned)
    except ValueError:
        pass
    return txt.strip()


# ============================================================================
# SpreadsheetML XML 2003
# ============================================================================


def _read_xml2003(path: Path) -> list[XlsSheet]:
    from lxml import etree

    root = etree.fromstring(path.read_bytes())
    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
    sheets: list[XlsSheet] = []
    for ws in root.findall(".//ss:Worksheet", ns):
        name = ws.get(f"{{{ns['ss']}}}Name") or "Sheet"
        rows: list[list[Cell]] = []
        max_cols = 0
        for row in ws.findall(".//ss:Row", ns):
            cells: list[Cell] = []
            for cell in row.findall("ss:Cell", ns):
                data = cell.find("ss:Data", ns)
                v: Cell = None
                if data is not None and data.text is not None:
                    raw_txt = data.text.strip()
                    t = data.get(f"{{{ns['ss']}}}Type", "String")
                    if t == "Number":
                        try:
                            v = float(raw_txt) if "." in raw_txt else int(raw_txt)
                        except ValueError:
                            v = raw_txt
                    elif t == "Boolean":
                        v = raw_txt in ("1", "true", "True")
                    else:
                        v = raw_txt or None
                cells.append(v)
            rows.append(cells)
            max_cols = max(max_cols, len(cells))
        sheets.append(XlsSheet(name=name, n_rows=len(rows), n_cols=max_cols, rows=rows))
    return sheets


# ============================================================================
# Helper comun
# ============================================================================


def _normalize(v: Any) -> Cell:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    return str(v).strip() or None
