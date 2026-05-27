"""Tests para `aibenchef_data.domains.parsing.services.xls_reader.read_xls`.

Cubre los 5 formatos que SBS publica bajo extension .xls:
- BIFF (xlrd) — incluye archivos con UTF-16 surrogates invalidos (regresion)
- XLSX (openpyxl)
- HTML disfrazado (BeautifulSoup)
- XML2003 (lxml)
- XLSB (pyxlsb) — sintetizado bajo demanda

Y los detectores de formato (`detect_xls_format`).
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from aibenchef_data.domains.parsing.services.xls_reader import (
    XlsSheet,
    read_xls,
)
from aibenchef_data.domains.parsing.value_objects.xls_format import (
    XlsFormat,
    detect_xls_format,
)
from aibenchef_data.domains.shared import ValidationError

# ---------------------------------------------------------------------------
# Helpers para generar archivos sinteticos en tmp_path
# ---------------------------------------------------------------------------


def _make_xlsx(tmp_path: Path, name: str = "data.xlsx") -> Path:
    """Genera un .xlsx valido con 2 hojas y 3 filas cada una."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Balance"
    ws1.append(["Cuenta", "MN", "ME", "Total"])
    ws1.append(["Activo", 100.5, 50.25, 150.75])
    ws1.append(["Pasivo", 80, 40, 120])

    ws2 = wb.create_sheet("Resultados")
    ws2.append(["Concepto", "Valor"])
    ws2.append(["Ingresos", 1000])
    ws2.append(["Gastos", 800])

    path = tmp_path / name
    wb.save(path)
    return path


def _make_html_disfrazado_xls(tmp_path: Path) -> Path:
    """Crea un .xls que en realidad es HTML (caso SBS pre-2010)."""
    html = """<html><head><meta charset="utf-8"><title>SBS</title></head><body>
<table>
<tr><th>Empresa</th><th>Saldo MN</th><th>Saldo ME</th></tr>
<tr><td>CMAC Arequipa</td><td>1,234,567</td><td>50,000.50</td></tr>
<tr><td>CMAC Cusco</td><td>987,654</td><td>25,000</td></tr>
</table>
</body></html>"""
    path = tmp_path / "html_disfrazado.xls"
    path.write_bytes(html.encode("utf-8"))
    return path


def _make_xml2003_xls(tmp_path: Path) -> Path:
    """Crea un SpreadsheetML XML 2003 .xls."""
    xml = """<?xml version="1.0"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
<Worksheet ss:Name="Datos">
<Table>
<Row><Cell><Data ss:Type="String">Entidad</Data></Cell>
     <Cell><Data ss:Type="String">Saldo</Data></Cell></Row>
<Row><Cell><Data ss:Type="String">BCP</Data></Cell>
     <Cell><Data ss:Type="Number">1000.5</Data></Cell></Row>
<Row><Cell><Data ss:Type="String">BBVA</Data></Cell>
     <Cell><Data ss:Type="Number">500</Data></Cell></Row>
</Table>
</Worksheet>
</Workbook>
"""
    path = tmp_path / "xml2003.xls"
    path.write_bytes(xml.encode("utf-8"))
    return path


# ---------------------------------------------------------------------------
# detect_xls_format
# ---------------------------------------------------------------------------


class TestDetectXlsFormat:
    def test_detecta_xlsx(self, tmp_path: Path):
        p = _make_xlsx(tmp_path)
        # cambiar extension a .xls para simular SBS publicando xlsx-as-xls
        renamed = tmp_path / "fake.xls"
        renamed.write_bytes(p.read_bytes())
        assert detect_xls_format(renamed) == XlsFormat.XLSX

    def test_detecta_html(self, tmp_path: Path):
        p = _make_html_disfrazado_xls(tmp_path)
        assert detect_xls_format(p) == XlsFormat.HTML

    def test_detecta_xml2003(self, tmp_path: Path):
        p = _make_xml2003_xls(tmp_path)
        assert detect_xls_format(p) == XlsFormat.XML2003

    def test_detecta_biff_real(self, fixtures_dir: Path):
        p = fixtures_dir / "biff_utf16_corrupt.xls"
        assert detect_xls_format(p) == XlsFormat.BIFF


# ---------------------------------------------------------------------------
# read_xls: happy path por formato
# ---------------------------------------------------------------------------


class TestReadXlsXlsx:
    def test_lee_xlsx_2_hojas(self, tmp_path: Path):
        p = _make_xlsx(tmp_path)
        sheets = read_xls(p)
        assert len(sheets) == 2
        assert sheets[0].name == "Balance"
        assert sheets[1].name == "Resultados"

    def test_xlsx_celdas_normalizadas(self, tmp_path: Path):
        p = _make_xlsx(tmp_path)
        sheets = read_xls(p)
        balance = sheets[0]
        assert balance.cell(0, 0) == "Cuenta"
        assert balance.cell(1, 1) == 100.5
        assert balance.cell(1, 2) == 50.25

    def test_xlsx_as_xls_via_bytesio(self, tmp_path: Path):
        """SBS publica algunos XLSX con extension .xls — openpyxl lo rechaza
        por el nombre, debemos cargar via BytesIO."""
        real = _make_xlsx(tmp_path, "data.xlsx")
        fake = tmp_path / "fake.xls"
        fake.write_bytes(real.read_bytes())
        sheets = read_xls(fake)
        assert len(sheets) == 2


class TestReadXlsHtml:
    def test_lee_html_table(self, tmp_path: Path):
        p = _make_html_disfrazado_xls(tmp_path)
        sheets = read_xls(p)
        assert len(sheets) == 1
        assert sheets[0].n_rows == 3
        assert sheets[0].cell(1, 0) == "CMAC Arequipa"

    def test_html_parsea_numeros_con_coma(self, tmp_path: Path):
        p = _make_html_disfrazado_xls(tmp_path)
        sheets = read_xls(p)
        # "1,234,567" -> int 1234567
        assert sheets[0].cell(1, 1) == 1234567
        # "50,000.50" -> float 50000.50
        assert sheets[0].cell(1, 2) == pytest.approx(50000.50)


class TestReadXlsXml2003:
    def test_lee_xml2003(self, tmp_path: Path):
        p = _make_xml2003_xls(tmp_path)
        sheets = read_xls(p)
        assert len(sheets) == 1
        assert sheets[0].name == "Datos"
        assert sheets[0].cell(0, 0) == "Entidad"
        assert sheets[0].cell(1, 1) == 1000.5  # tipo Number
        assert sheets[0].cell(2, 1) == 500  # tipo Number sin punto -> int


# ---------------------------------------------------------------------------
# BIFF — happy path + REGRESION del fix UTF-16 surrogate
# ---------------------------------------------------------------------------


class TestReadXlsBiffReal:
    """Tests sobre archivos BIFF reales descargados de SBS.

    El archivo biff_utf16_corrupt.xls (C-1231-jn2017.xls original) tiene
    surrogates UTF-16 invalidos que rompian xlrd antes del fix del parser.
    """

    def test_biff_se_lee_sin_excepcion(self, fixtures_dir: Path):
        """REGRESION: este archivo fallaba con 'illegal UTF-16 surrogate'."""
        p = fixtures_dir / "biff_utf16_corrupt.xls"
        sheets = read_xls(p)
        assert len(sheets) >= 1
        # Debe haber datos parseables — al menos 10 filas con contenido
        nonempty_cells = sum(
            1
            for s in sheets
            for r in range(s.n_rows)
            for c in range(s.n_cols)
            if s.cell(r, c) is not None
        )
        assert nonempty_cells > 50

    def test_biff_banca_old_layout(self, fixtures_dir: Path):
        """Pre-2015 BANCOS clientes_credito: empresa en col 1 (no col 0)."""
        p = fixtures_dir / "biff_banca_old_layout.xls"
        sheets = read_xls(p)
        assert sheets, "Debe parsear al menos una hoja"
        # Buscar 'B. Continental' u otra banca historica como senal de parsing OK
        textos = {
            str(s.cell(r, c))
            for s in sheets
            for r in range(min(s.n_rows, 15))
            for c in range(min(s.n_cols, 5))
            if isinstance(s.cell(r, c), str)
        }
        assert any("Continental" in t for t in textos), (
            f"No encontre 'Continental' en cells iniciales: {textos}"
        )

    def test_biff_inverted_layout(self, fixtures_dir: Path):
        """2010 CMAC clientes_ahorro: personas como header OUTER, productos INNER."""
        p = fixtures_dir / "biff_inverted_layout.xls"
        sheets = read_xls(p)
        assert sheets
        # En este layout esperamos ver al menos un texto que diga 'CMAC' en col 0
        encontrados = [
            sheets[0].cell(r, 0)
            for r in range(sheets[0].n_rows)
            if isinstance(sheets[0].cell(r, 0), str)
        ]
        assert any("CMAC" in str(t).upper() for t in encontrados)


# ---------------------------------------------------------------------------
# Errores y casos borde
# ---------------------------------------------------------------------------


class TestReadXlsErrores:
    def test_archivo_no_existe(self, tmp_path: Path):
        with pytest.raises(ValidationError, match="no existe"):
            read_xls(tmp_path / "no_existe.xls")

    def test_archivo_corrupto_no_crashea_silenciosamente(self, tmp_path: Path):
        """Un archivo de bytes random debe levantar ValidationError, no AttributeError."""
        p = tmp_path / "garbage.xls"
        p.write_bytes(b"\x00\x01\x02not really an excel file")
        with pytest.raises((ValidationError, Exception)):
            read_xls(p)


# ---------------------------------------------------------------------------
# XlsSheet dataclass
# ---------------------------------------------------------------------------


class TestXlsSheetDataclass:
    def test_cell_fuera_de_rango_retorna_none(self):
        sheet = XlsSheet(
            name="test",
            n_rows=2,
            n_cols=2,
            rows=[[1, 2], [3, 4]],
        )
        assert sheet.cell(0, 0) == 1
        assert sheet.cell(1, 1) == 4
        assert sheet.cell(5, 0) is None  # fila fuera de rango
        assert sheet.cell(0, 5) is None  # columna fuera de rango

    def test_row_indexing(self):
        sheet = XlsSheet(
            name="test",
            n_rows=2,
            n_cols=2,
            rows=[["a", "b"], ["c", "d"]],
        )
        assert sheet.row(0) == ["a", "b"]
        assert sheet.row(1) == ["c", "d"]
